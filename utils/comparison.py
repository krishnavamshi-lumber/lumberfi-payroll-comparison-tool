"""
Comparison utilities for PDFs and CSVs.
"""

from __future__ import annotations

import difflib
import io
import os
import sys

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("Missing dependency. Run:  pip install pymupdf")


# ── Internal helpers ───────────────────────────────────────────────────────────

def _get_words(page: fitz.Page) -> list:
    return [(w[0], w[1], w[2], w[3], w[4]) for w in page.get_text("words")]


def _diff_word_lists(words_a: list, words_b: list) -> tuple[set, set]:
    seq_a = [w[4] for w in words_a]
    seq_b = [w[4] for w in words_b]
    sm = difflib.SequenceMatcher(None, seq_a, seq_b, autojunk=False)
    deleted: set[int]  = set()
    inserted: set[int] = set()
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag in ("delete", "replace"):
            deleted.update(range(i1, i2))
        if tag in ("insert", "replace"):
            inserted.update(range(j1, j2))
    return deleted, inserted


def _build_combined_page(
    out_doc: fitz.Document,
    doc_a: fitz.Document,
    doc_b: fitz.Document,
    page_num: int,
    total_pages: int,
) -> None:
    has_a = page_num < len(doc_a)
    has_b = page_num < len(doc_b)

    ref_page = doc_a[page_num] if has_a else doc_b[page_num]
    pw = ref_page.rect.width
    ph = ref_page.rect.height

    DIVIDER = 2
    LABEL_H = 22
    canvas_w = pw * 2 + DIVIDER
    canvas_h = ph + LABEL_H

    new_page = out_doc.new_page(width=canvas_w, height=canvas_h)
    new_page.draw_rect(
        fitz.Rect(0, 0, canvas_w, canvas_h),
        color=None, fill=(0.96, 0.95, 0.93),
    )

    font_size = 9
    label_y   = LABEL_H - 5
    new_page.insert_text(
        (8, label_y),
        f"TRUTH  (page {page_num + 1}/{total_pages})  — removed text highlighted red",
        fontsize=font_size, color=(0.5, 0.1, 0.1),
    )
    new_page.insert_text(
        (pw + DIVIDER + 8, label_y),
        f"COMPARE  (page {page_num + 1}/{total_pages})  — added text highlighted green",
        fontsize=font_size, color=(0.1, 0.4, 0.1),
    )

    new_page.draw_line(
        (pw + DIVIDER / 2, 0),
        (pw + DIVIDER / 2, canvas_h),
        color=(0.7, 0.7, 0.7), width=1,
    )

    left_rect  = fitz.Rect(0,            LABEL_H, pw,       canvas_h)
    right_rect = fitz.Rect(pw + DIVIDER, LABEL_H, canvas_w, canvas_h)

    if has_a:
        new_page.show_pdf_page(left_rect,  doc_a, page_num)
    if has_b:
        new_page.show_pdf_page(right_rect, doc_b, page_num)

    words_a = _get_words(doc_a[page_num]) if has_a else []
    words_b = _get_words(doc_b[page_num]) if has_b else []
    deleted, inserted = _diff_word_lists(words_a, words_b)

    for idx in deleted:
        if idx < len(words_a):
            w = words_a[idx]
            rect = fitz.Rect(w[0], w[1] + LABEL_H, w[2], w[3] + LABEL_H)
            new_page.draw_rect(rect, color=None, fill=(1.0, 0.2, 0.2), fill_opacity=0.3, overlay=True)

    for idx in inserted:
        if idx < len(words_b):
            w = words_b[idx]
            rect = fitz.Rect(
                w[0] + pw + DIVIDER, w[1] + LABEL_H,
                w[2] + pw + DIVIDER, w[3] + LABEL_H,
            )
            new_page.draw_rect(rect, color=None, fill=(0.2, 0.85, 0.2), fill_opacity=0.3, overlay=True)


def _build_summary_page(
    out_doc: fitz.Document,
    doc_a: fitz.Document,
    doc_b: fitz.Document,
    previous_label: str,
    current_label: str,
    project_name: str,
    report_type: str,
) -> None:
    total_del = total_ins = total_eq = 0
    max_pages = max(len(doc_a), len(doc_b))

    for i in range(max_pages):
        wa = _get_words(doc_a[i]) if i < len(doc_a) else []
        wb = _get_words(doc_b[i]) if i < len(doc_b) else []
        deleted, inserted = _diff_word_lists(wa, wb)
        sm = difflib.SequenceMatcher(
            None, [w[4] for w in wa], [w[4] for w in wb], autojunk=False,
        )
        eq = sum(i2 - i1 for tag, i1, i2, _, _ in sm.get_opcodes() if tag == "equal")
        total_del += len(deleted)
        total_ins += len(inserted)
        total_eq  += eq

    total_words = total_del + total_eq
    pct = round((total_del + total_ins) / max(total_words + total_ins, 1) * 100, 1)

    pg = out_doc.new_page(width=595, height=842)
    pg.draw_rect(fitz.Rect(0, 0, 595, 60), color=None, fill=(0.1, 0.1, 0.1))
    pg.insert_text(
        (24, 38), "PDF COMPARISON",
        fontsize=16, color=(1, 1, 1), fontname="helv",
    )

    y = [90]

    def line(text, x=24, size=11, color=(0.15, 0.15, 0.15)):
        pg.insert_text((x, y[0]), text, fontsize=size, color=color, fontname="helv")
        y[0] += size + 8

    line(f"Project: {project_name}  |  Report type: {report_type}", size=10, color=(0.2, 0.2, 0.6))
    y[0] += 6

    line("Files compared", size=9, color=(0.5, 0.5, 0.5))
    line(f"  Truth (A) :  {previous_label}", size=10)
    line(f"  Compare  (B) :  {current_label}",  size=10)
    y[0] += 10

    line("Summary", size=9, color=(0.5, 0.5, 0.5))
    line(f"  Pages compared   :  {max_pages}",    size=10)
    line(f"  Words added      :  +{total_ins}",   size=10, color=(0.1, 0.45, 0.1))
    line(f"  Words removed    :  -{total_del}",   size=10, color=(0.6, 0.1, 0.1))
    line(f"  Words unchanged  :  {total_eq}",     size=10)
    line(f"  % changed        :  {pct}%",         size=10)
    y[0] += 14

    line("Highlight legend", size=9, color=(0.5, 0.5, 0.5))
    pg.draw_rect(fitz.Rect(24, y[0] - 10, 44, y[0] + 2), color=None, fill=(1.0, 0.2, 0.2), fill_opacity=0.5)
    pg.insert_text((50, y[0]), "Red   = text removed from Truth report",
                   fontsize=10, color=(0.15, 0.15, 0.15), fontname="helv")
    y[0] += 18

    pg.draw_rect(fitz.Rect(24, y[0] - 10, 44, y[0] + 2), color=None, fill=(0.2, 0.85, 0.2), fill_opacity=0.5)
    pg.insert_text((50, y[0]), "Green = text added in Compare report",
                   fontsize=10, color=(0.15, 0.15, 0.15), fontname="helv")
    y[0] += 30

    line("Side-by-side comparison follows on the next pages.", size=9, color=(0.5, 0.5, 0.5))


# ── Public API ─────────────────────────────────────────────────────────────────

def _get_file_extension(filename: str) -> str:
    return os.path.splitext(filename or "")[1].strip().lower()


def compare_pdfs(
    truth_pdf_bytes: bytes,
    compare_pdf_bytes: bytes,
    project_name: str = "Project",
    report_type: str = "Report",
    truth_label: str = "Truth",
    compare_label: str = "Compare",
) -> bytes:
    """
    Compare two PDFs and return a highlighted diff PDF as bytes.

    Args:
        truth_pdf_bytes: PDF bytes from the truth file (File A — left column).
        compare_pdf_bytes:  PDF bytes from the compare file  (File B — right column).
        project_name:       Human-readable project name shown on the summary page.
        report_type:        Type of report — shown on the summary page.
        truth_label:     Column A label.
        compare_label:      Column B label.

    Returns:
        Bytes of the generated diff PDF (summary page + side-by-side pages).
    """
    doc_a = fitz.open(stream=truth_pdf_bytes, filetype="pdf")
    doc_b = fitz.open(stream=compare_pdf_bytes,  filetype="pdf")
    out_doc = fitz.open()

    try:
        _build_summary_page(
            out_doc, doc_a, doc_b,
            truth_label, compare_label,
            project_name, report_type,
        )
        max_pages = max(len(doc_a), len(doc_b))
        for i in range(max_pages):
            _build_combined_page(out_doc, doc_a, doc_b, i, max_pages)

        buf = io.BytesIO()
        out_doc.save(buf)
        return buf.getvalue()
    finally:
        out_doc.close()
        doc_a.close()
        doc_b.close()


_WORKER_COMP_KEY_COLUMNS = [
    "SSN",
    "Date Worked",
    "Project Code",
    "Cost Code",
    "Task Code",
    "Regular Hours",
    "Overtime Hours",
    "Double Overtime Hours",
]


def compare_worker_comp_csvs(
    truth_csv_bytes: bytes,
    compare_csv_bytes: bytes,
    truth_label: str = "Truth",
    compare_label: str = "Compare",
) -> bytes:
    """
    Compare two Worker Compensation CSV files using key-based row matching.

    Rows are matched by (SSN, Date Worked, Project Code, Cost Code, Task Code,
    Regular Hours, Overtime Hours, Double Overtime Hours) rather than position,
    so inserted or reordered rows do not cascade false CHANGEs.

    Duplicate keys are paired first-to-first. Extra rows on either side are
    reported as ADDED or REMOVED.
    """
    try:
        import pandas as pd
        from collections import defaultdict
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError as exc:
        raise ImportError(
            f"Missing dependency for tabular comparison: {exc}. "
            "Run:  pip install pandas openpyxl"
        ) from exc

    RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def _read_csv(file_bytes: bytes):
        return pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False).fillna("")

    df_a = _read_csv(truth_csv_bytes)
    df_b = _read_csv(compare_csv_bytes)

    columns = list(dict.fromkeys(list(df_a.columns) + list(df_b.columns)))
    df_a = df_a.reindex(columns=columns, fill_value="")
    df_b = df_b.reindex(columns=columns, fill_value="")

    active_key_cols = [c for c in _WORKER_COMP_KEY_COLUMNS if c in columns]

    def _row_key(row: dict) -> str:
        return "|".join(str(row.get(col, "")).strip() for col in active_key_cols)

    def _group_by_key(df):
        groups: dict[str, list[dict]] = defaultdict(list)
        key_order: list[str] = []
        for _, row in df.iterrows():
            k = _row_key(row.to_dict())
            if k not in groups:
                key_order.append(k)
            groups[k].append(row.to_dict())
        return groups, key_order

    groups_a, order_a = _group_by_key(df_a)
    groups_b, order_b = _group_by_key(df_b)

    seen: set[str] = set(order_a)
    all_keys = list(order_a)
    for k in order_b:
        if k not in seen:
            all_keys.append(k)
            seen.add(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "Worker Comp Comparison"

    headers = ["Row", "Status"]
    for col in columns:
        headers.append(f"{truth_label}  {col}")
        headers.append(f"{compare_label}  {col}")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    output_row_num = 0
    empty_row = {col: "" for col in columns}

    for key in all_keys:
        rows_a = groups_a.get(key, [])
        rows_b = groups_b.get(key, [])
        max_count = max(len(rows_a), len(rows_b))

        for i in range(max_count):
            row_a = rows_a[i] if i < len(rows_a) else empty_row
            row_b = rows_b[i] if i < len(rows_b) else empty_row

            output_row_num += 1

            if i >= len(rows_a):
                status = "ADDED"
            elif i >= len(rows_b):
                status = "REMOVED"
            elif any(row_a.get(col, "") != row_b.get(col, "") for col in columns):
                status = "CHANGED"
            else:
                status = "UNCHANGED"

            row_data = [output_row_num, status]
            changed_flags = []
            for col in columns:
                row_data.append(row_a.get(col, ""))
                row_data.append(row_b.get(col, ""))
                changed_flags.append(row_a.get(col, "") != row_b.get(col, ""))

            ws.append(row_data)
            excel_row = ws.max_row

            if status == "ADDED":
                for col_idx in range(3, len(headers) + 1):
                    ws.cell(excel_row, col_idx).fill = GREEN_FILL
            elif status == "REMOVED":
                for col_idx in range(3, len(headers) + 1):
                    ws.cell(excel_row, col_idx).fill = RED_FILL
            elif status == "CHANGED":
                col_pointer = 3
                for changed in changed_flags:
                    if changed:
                        ws.cell(excel_row, col_pointer).fill = YELLOW_FILL
                        ws.cell(excel_row, col_pointer + 1).fill = YELLOW_FILL
                    col_pointer += 2

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def compare_any_files(
    truth_bytes: bytes,
    compare_bytes: bytes,
    truth_filename: str,
    compare_filename: str,
    project_name: str = "Project",
    report_type: str = "Report",
    truth_label: str = "Truth",
    compare_label: str = "Compare",
) -> tuple[bytes, str, str]:
    """
    Select the correct comparison function based on file extension.

    Returns:
        A tuple of (output_bytes, output_extension, mime_type).
    """
    truth_ext = _get_file_extension(truth_filename)
    compare_ext = _get_file_extension(compare_filename)

    if truth_ext != compare_ext:
        raise ValueError(
            f"Truth and compare files must have the same extension: "
            f"{truth_ext} != {compare_ext}"
        )

    if truth_ext == ".pdf":
        return (
            compare_pdfs(
                truth_bytes,
                compare_bytes,
                project_name=project_name,
                report_type=report_type,
                truth_label=truth_label,
                compare_label=compare_label,
            ),
            "pdf",
            "application/pdf",
        )

    if truth_ext in (".csv", ".xlsx", ".xls"):
        compare_fn = (
            compare_worker_comp_csvs
            if report_type == "Worker Compensation"
            else compare_csvs
        )
        return (
            compare_fn(
                truth_bytes,
                compare_bytes,
                truth_label=truth_label,
                compare_label=compare_label,
            ),
            "xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    raise ValueError(f"Unsupported file type for comparison: {truth_ext}")


def _read_tabular_file(file_bytes: bytes) -> dict[str, object]:
    import pandas as pd

    if file_bytes.startswith(b'PK\x03\x04'):
        sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, dtype=str)
        return {name: df.fillna("") for name, df in sheets.items()}

    df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False).fillna("")
    return {"Sheet1": df}


def _normalize_sheet_title(name: str, existing: set[str]) -> str:
    safe = name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_')
    candidate = safe or 'Sheet'
    suffix = 1
    while candidate in existing:
        candidate = f"{safe[:28]}_{suffix}" if len(safe) > 28 else f"{safe}_{suffix}"
        suffix += 1
    existing.add(candidate)
    return candidate


def compare_csvs(
    truth_csv_bytes: bytes,
    compare_csv_bytes: bytes,
    truth_label: str = "Truth",
    compare_label: str = "Compare",
) -> bytes:
    """
    Compare two tabular files (CSV or Excel) and return a highlighted Excel workbook as bytes.

    For Excel files, all sheets are compared by sheet name. If both files contain
    a sheet named "Sheet1" and "Sheet2", the function compares those sheets
    pairwise. If one workbook has extra sheets, those sheets are compared with an
    empty sheet.

    Rows are compared by position. Columns are aligned by name.
    Colour coding:
      Red    = row removed from truth
      Green  = row added to compare
      Yellow = row present in both but with changed cell values

    Args:
        truth_csv_bytes: Raw bytes of the truth file (CSV or Excel).
        compare_csv_bytes: Raw bytes of the compare file (CSV or Excel).
        truth_label:     Label used in header for the A (truth) columns.
        compare_label:      Label used in header for the B (compare) columns.

    Returns:
        Bytes of an .xlsx workbook with highlighted differences.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError as exc:
        raise ImportError(
            f"Missing dependency for tabular comparison: {exc}. "
            "Run:  pip install pandas openpyxl"
        ) from exc

    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    tabs_a = _read_tabular_file(truth_csv_bytes)
    tabs_b = _read_tabular_file(compare_csv_bytes)
    sheet_names = list(dict.fromkeys(list(tabs_a) + list(tabs_b)))

    wb = Workbook()
    existing_titles: set[str] = set()
    first_sheet = True

    for sheet_name in sheet_names:
        df_a = tabs_a.get(sheet_name, pd.DataFrame())
        df_b = tabs_b.get(sheet_name, pd.DataFrame())

        columns = sorted(set(df_a.columns) | set(df_b.columns))
        df_a = df_a.reindex(columns=columns, fill_value="")
        df_b = df_b.reindex(columns=columns, fill_value="")

        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()

        ws.title = _normalize_sheet_title(sheet_name, existing_titles)

        headers = ["Row", "Status"]
        for col in columns:
            headers.append(f"{truth_label}  {col}")
            headers.append(f"{compare_label}  {col}")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)

        max_rows = max(len(df_a), len(df_b))

        for row_index in range(max_rows):
            row_a = df_a.iloc[row_index].to_dict() if row_index < len(df_a) else {col: "" for col in columns}
            row_b = df_b.iloc[row_index].to_dict() if row_index < len(df_b) else {col: "" for col in columns}

            if row_index >= len(df_a):
                status = "ADDED"
            elif row_index >= len(df_b):
                status = "REMOVED"
            else:
                status = "CHANGED" if any(row_a[col] != row_b[col] for col in columns) else "UNCHANGED"

            row = [row_index + 1, status]
            changed_flags = []
            for col in columns:
                row.append(row_a.get(col, ""))
                row.append(row_b.get(col, ""))
                changed_flags.append(row_a.get(col, "") != row_b.get(col, ""))

            ws.append(row)
            excel_row = ws.max_row

            if status == "ADDED":
                for col_idx in range(3, len(headers) + 1):
                    ws.cell(excel_row, col_idx).fill = GREEN_FILL
            elif status == "REMOVED":
                for col_idx in range(3, len(headers) + 1):
                    ws.cell(excel_row, col_idx).fill = RED_FILL
            else:
                col_pointer = 3
                for changed in changed_flags:
                    if changed:
                        ws.cell(excel_row, col_pointer).fill = YELLOW_FILL
                        ws.cell(excel_row, col_pointer + 1).fill = YELLOW_FILL
                    col_pointer += 2

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# """
# Structured comparison utilities for PDFs and CSVs.

# PDF comparison:
# - Parses payroll employee sections instead of raw word-stream diffs
# - Matches employees using emp_key = normalized_name|identity_number
# - Highlights:
#     RED    = employee removed
#     GREEN  = employee added
#     YELLOW = employee present but changed

# CSV comparison:
# - Existing emp_key comparison retained
# """

# from __future__ import annotations

# import difflib
# import io
# import re
# import sys
# from dataclasses import dataclass
# from typing import Dict, List, Tuple

# try:
#     import fitz  # PyMuPDF
# except ImportError:
#     sys.exit("Missing dependency. Run: pip install pymupdf")


# # =============================================================================
# # PDF EMPLOYEE STRUCTURES
# # =============================================================================

# @dataclass
# class EmployeeRecord:
#     emp_key: str
#     name: str
#     identity: str
#     page_num: int
#     rect: fitz.Rect
#     raw_text: str
#     normalized_text: str


# # =============================================================================
# # NORMALIZATION HELPERS
# # =============================================================================

# def _normalize_text(text: str) -> str:
#     text = text.replace("\n", " ")
#     text = text.replace("\u00ad", "")
#     text = text.replace("￾", "")
#     text = re.sub(r"\s+", " ", text)
#     return text.strip()


# def _normalize_name(name: str) -> str:
#     """
#     Generic name normalization.

#     Handles:
#     - weird spacing
#     - OCR artifacts
#     - unicode junk
#     - inconsistent separators
#     """

#     name = _normalize_text(name).lower()

#     # remove invisible / junk unicode chars
#     name = re.sub(r"[\u00ad\ufffe\uffff]", "", name)

#     # collapse multiple spaces
#     name = re.sub(r"\s+", " ", name)

#     # keep only reasonable name chars
#     name = re.sub(r"[^a-z0-9,\- ]", "", name)

#     return name.strip()


# def _clean_identity(identity: str) -> str:
#     identity = re.sub(r"[^\d\-]", "", identity)
#     return identity.strip()


# # =============================================================================
# # EMPLOYEE EXTRACTION
# # =============================================================================

# SSN_RE = re.compile(r"\d{3}-\d{2}-\d{4}")


# def _extract_employee_blocks(page: fitz.Page) -> List[Tuple[fitz.Rect, str]]:
#     """
#     Extract visually grouped text blocks.
#     """
#     blocks = page.get_text("blocks")

#     extracted = []

#     for block in blocks:
#         x0, y0, x1, y1, text, *_ = block

#         text = _normalize_text(text)

#         if not text:
#             continue

#         if not SSN_RE.search(text):
#             continue

#         extracted.append((fitz.Rect(x0, y0, x1, y1), text))

#     return extracted


# def _extract_name_and_ssn(text: str) -> Tuple[str, str]:
#     ssn_match = SSN_RE.search(text)

#     if not ssn_match:
#         return "", ""

#     ssn = ssn_match.group(0)

#     before_ssn = text[:ssn_match.start()].strip()

#     lines = [l.strip() for l in before_ssn.splitlines() if l.strip()]

#     if not lines:
#         return "", ssn

#     candidate = lines[-1]

#     candidate = _normalize_text(candidate)

#     return candidate, ssn


# def _build_emp_key(name: str, identity: str) -> str:
#     compact_name = re.sub(r"[^a-z]", "", name.lower())
#     clean_ssn = re.sub(r"[^\d]", "", identity)

#     return f"{compact_name}|{clean_ssn}"


# def _extract_employee_records(doc: fitz.Document) -> Dict[str, EmployeeRecord]:
#     """
#     Parse employee records from payroll PDF.
#     """

#     records: Dict[str, EmployeeRecord] = {}

#     for page_num in range(len(doc)):
#         page = doc[page_num]

#         for rect, text in _extract_employee_blocks(page):

#             name, identity = _extract_name_and_ssn(text)

#             if not name or not identity:
#                 continue

#             emp_key = _build_emp_key(name, identity)

#             normalized_text = _normalize_text(text).lower()

#             if emp_key in records:
#                 existing = records[emp_key]

#                 merged_text = (
#                     existing.normalized_text
#                     + "\n"
#                     + normalized_text
#                 )

#                 merged_rect = fitz.Rect(
#                     min(existing.rect.x0, rect.x0),
#                     min(existing.rect.y0, rect.y0),
#                     max(existing.rect.x1, rect.x1),
#                     max(existing.rect.y1, rect.y1),
#                 )

#                 records[emp_key] = EmployeeRecord(
#                     emp_key=emp_key,
#                     name=name,
#                     identity=identity,
#                     page_num=page_num,
#                     rect=merged_rect,
#                     raw_text=existing.raw_text + "\n" + text,
#                     normalized_text=merged_text,
#                 )
#             else:
#                 records[emp_key] = EmployeeRecord(
#                     emp_key=emp_key,
#                     name=name,
#                     identity=identity,
#                     page_num=page_num,
#                     rect=rect,
#                     raw_text=text,
#                     normalized_text=normalized_text,
#                 )

#     return records


# # =============================================================================
# # PDF VISUAL PAGE BUILDER
# # =============================================================================

# def _draw_highlight(
#     page: fitz.Page,
#     rect: fitz.Rect,
#     color: Tuple[float, float, float],
# ):
#     pad = 2

#     expanded = fitz.Rect(
#         rect.x0 - pad,
#         rect.y0 - pad,
#         rect.x1 + pad,
#         rect.y1 + pad,
#     )

#     page.draw_rect(
#         expanded,
#         color=None,
#         fill=color,
#         fill_opacity=0.28,
#         overlay=True,
#     )


# def _record_changed(a: EmployeeRecord, b: EmployeeRecord) -> bool:
#     return a.normalized_text != b.normalized_text


# def _build_combined_page(
#     out_doc: fitz.Document,
#     doc_a: fitz.Document,
#     doc_b: fitz.Document,
#     page_num: int,
#     total_pages: int,
#     a_records: Dict[str, EmployeeRecord],
#     b_records: Dict[str, EmployeeRecord],
# ):
#     has_a = page_num < len(doc_a)
#     has_b = page_num < len(doc_b)

#     ref_page = doc_a[page_num] if has_a else doc_b[page_num]

#     pw = ref_page.rect.width
#     ph = ref_page.rect.height

#     DIVIDER = 2
#     LABEL_H = 24

#     canvas_w = pw * 2 + DIVIDER
#     canvas_h = ph + LABEL_H

#     new_page = out_doc.new_page(width=canvas_w, height=canvas_h)

#     new_page.draw_rect(
#         fitz.Rect(0, 0, canvas_w, canvas_h),
#         color=None,
#         fill=(0.96, 0.95, 0.93),
#     )

#     new_page.insert_text(
#         (8, 16),
#         f"TRUTH  (page {page_num + 1}/{total_pages})",
#         fontsize=10,
#         color=(0.5, 0.1, 0.1),
#     )

#     new_page.insert_text(
#         (pw + DIVIDER + 8, 16),
#         f"COMPARE  (page {page_num + 1}/{total_pages})",
#         fontsize=10,
#         color=(0.1, 0.4, 0.1),
#     )

#     new_page.draw_line(
#         (pw + DIVIDER / 2, 0),
#         (pw + DIVIDER / 2, canvas_h),
#         color=(0.7, 0.7, 0.7),
#         width=1,
#     )

#     left_rect = fitz.Rect(0, LABEL_H, pw, canvas_h)
#     right_rect = fitz.Rect(pw + DIVIDER, LABEL_H, canvas_w, canvas_h)

#     if has_a:
#         new_page.show_pdf_page(left_rect, doc_a, page_num)

#     if has_b:
#         new_page.show_pdf_page(right_rect, doc_b, page_num)

#     # -------------------------------------------------------------------------
#     # Highlight employee changes
#     # -------------------------------------------------------------------------

#     all_keys = set(a_records) | set(b_records)

#     for key in all_keys:

#         rec_a = a_records.get(key)
#         rec_b = b_records.get(key)

#         # REMOVED
#         if rec_a and not rec_b:
#             if rec_a.page_num == page_num:

#                 rect = fitz.Rect(
#                     rec_a.rect.x0,
#                     rec_a.rect.y0 + LABEL_H,
#                     rec_a.rect.x1,
#                     rec_a.rect.y1 + LABEL_H,
#                 )

#                 _draw_highlight(
#                     new_page,
#                     rect,
#                     (1.0, 0.2, 0.2),
#                 )

#         # ADDED
#         elif rec_b and not rec_a:
#             if rec_b.page_num == page_num:

#                 rect = fitz.Rect(
#                     rec_b.rect.x0 + pw + DIVIDER,
#                     rec_b.rect.y0 + LABEL_H,
#                     rec_b.rect.x1 + pw + DIVIDER,
#                     rec_b.rect.y1 + LABEL_H,
#                 )

#                 _draw_highlight(
#                     new_page,
#                     rect,
#                     (0.2, 0.85, 0.2),
#                 )

#         # CHANGED
#         elif rec_a and rec_b:
#             if _record_changed(rec_a, rec_b):

#                 if rec_a.page_num == page_num:

#                     rect = fitz.Rect(
#                         rec_a.rect.x0,
#                         rec_a.rect.y0 + LABEL_H,
#                         rec_a.rect.x1,
#                         rec_a.rect.y1 + LABEL_H,
#                     )

#                     _draw_highlight(
#                         new_page,
#                         rect,
#                         (1.0, 0.9, 0.2),
#                     )

#                 if rec_b.page_num == page_num:

#                     rect = fitz.Rect(
#                         rec_b.rect.x0 + pw + DIVIDER,
#                         rec_b.rect.y0 + LABEL_H,
#                         rec_b.rect.x1 + pw + DIVIDER,
#                         rec_b.rect.y1 + LABEL_H,
#                     )

#                     _draw_highlight(
#                         new_page,
#                         rect,
#                         (1.0, 0.9, 0.2),
#                     )


# # =============================================================================
# # SUMMARY PAGE
# # =============================================================================

# def _build_summary_page(
#     out_doc: fitz.Document,
#     previous_label: str,
#     current_label: str,
#     project_name: str,
#     report_type: str,
#     a_records: Dict[str, EmployeeRecord],
#     b_records: Dict[str, EmployeeRecord],
# ):
#     added = 0
#     removed = 0
#     changed = 0
#     unchanged = 0

#     all_keys = set(a_records) | set(b_records)

#     for key in all_keys:

#         rec_a = a_records.get(key)
#         rec_b = b_records.get(key)

#         if rec_a and not rec_b:
#             removed += 1

#         elif rec_b and not rec_a:
#             added += 1

#         elif rec_a and rec_b:
#             if _record_changed(rec_a, rec_b):
#                 changed += 1
#             else:
#                 unchanged += 1

#     pg = out_doc.new_page(width=595, height=842)

#     pg.draw_rect(
#         fitz.Rect(0, 0, 595, 60),
#         color=None,
#         fill=(0.1, 0.1, 0.1),
#     )

#     pg.insert_text(
#         (24, 38),
#         "STRUCTURED PDF COMPARISON",
#         fontsize=16,
#         color=(1, 1, 1),
#     )

#     y = 100

#     lines = [
#         f"Project: {project_name}",
#         f"Report type: {report_type}",
#         "",
#         f"Truth file   : {previous_label}",
#         f"Compare file : {current_label}",
#         "",
#         f"Employees added      : {added}",
#         f"Employees removed    : {removed}",
#         f"Employees changed    : {changed}",
#         f"Employees unchanged  : {unchanged}",
#         "",
#         "Legend:",
#         "RED    = removed employee",
#         "GREEN  = added employee",
#         "YELLOW = changed employee",
#     ]

#     for line in lines:
#         pg.insert_text(
#             (24, y),
#             line,
#             fontsize=11,
#             color=(0.15, 0.15, 0.15),
#         )
#         y += 22


# # =============================================================================
# # PUBLIC PDF API
# # =============================================================================

# def compare_pdfs(
#     truth_pdf_bytes: bytes,
#     compare_pdf_bytes: bytes,
#     project_name: str = "Project",
#     report_type: str = "Report",
#     truth_label: str = "Truth",
#     compare_label: str = "Compare",
# ) -> bytes:
#     """
#     Structured payroll PDF comparison.

#     Uses employee-level matching instead of raw word diffs.
#     """

#     doc_a = fitz.open(stream=truth_pdf_bytes, filetype="pdf")
#     doc_b = fitz.open(stream=compare_pdf_bytes, filetype="pdf")

#     out_doc = fitz.open()

#     try:

#         a_records = _extract_employee_records(doc_a)
#         b_records = _extract_employee_records(doc_b)

#         _build_summary_page(
#             out_doc,
#             truth_label,
#             compare_label,
#             project_name,
#             report_type,
#             a_records,
#             b_records,
#         )

#         max_pages = max(len(doc_a), len(doc_b))

#         for page_num in range(max_pages):
#             _build_combined_page(
#                 out_doc,
#                 doc_a,
#                 doc_b,
#                 page_num,
#                 max_pages,
#                 a_records,
#                 b_records,
#             )

#         buf = io.BytesIO()

#         out_doc.save(buf)

#         return buf.getvalue()

#     finally:
#         out_doc.close()
#         doc_a.close()
#         doc_b.close()


# # =============================================================================
# # CSV COMPARISON
# # =============================================================================

# def compare_csvs(
#     truth_csv_bytes: bytes,
#     compare_csv_bytes: bytes,
#     truth_label: str = "Truth",
#     compare_label: str = "Compare",
# ) -> bytes:

#     try:
#         import pandas as pd
#         from openpyxl import Workbook
#         from openpyxl.styles import PatternFill, Font, Alignment

#     except ImportError as exc:
#         raise ImportError(
#             f"Missing dependency for CSV comparison: {exc}. "
#             "Run: pip install pandas openpyxl"
#         ) from exc

#     RED_FILL = PatternFill(
#         start_color="FFC7CE",
#         end_color="FFC7CE",
#         fill_type="solid",
#     )

#     GREEN_FILL = PatternFill(
#         start_color="C6EFCE",
#         end_color="C6EFCE",
#         fill_type="solid",
#     )

#     YELLOW_FILL = PatternFill(
#         start_color="FFEB9C",
#         end_color="FFEB9C",
#         fill_type="solid",
#     )

#     df_a = pd.read_csv(
#         io.BytesIO(truth_csv_bytes),
#         dtype=str,
#     ).fillna("")

#     df_b = pd.read_csv(
#         io.BytesIO(compare_csv_bytes),
#         dtype=str,
#     ).fillna("")

#     df_a["_emp_key"] = (
#         df_a["first_name"].str.strip()
#         + "|"
#         + df_a["last_name"].str.strip()
#         + "|"
#         + df_a["class_code"].str.strip()
#     )

#     df_b["_emp_key"] = (
#         df_b["first_name"].str.strip()
#         + "|"
#         + df_b["last_name"].str.strip()
#         + "|"
#         + df_b["class_code"].str.strip()
#     )

#     a_map = df_a.set_index("_emp_key").to_dict(orient="index")
#     b_map = df_b.set_index("_emp_key").to_dict(orient="index")

#     columns = [c for c in df_a.columns if c != "_emp_key"]

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "CSV Comparison"

#     headers = ["Employee", "Status"]

#     for col in columns:
#         headers.append(f"{truth_label} {col}")
#         headers.append(f"{compare_label} {col}")

#     ws.append(headers)

#     for cell in ws[1]:
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(wrap_text=True)

#     all_keys = sorted(set(a_map) | set(b_map))

#     for key in all_keys:

#         row_a = a_map.get(key, {})
#         row_b = b_map.get(key, {})

#         base_row = [key]

#         # REMOVED
#         if row_a and not row_b:

#             base_row.append("REMOVED")

#             for col in columns:
#                 base_row.extend([row_a.get(col, ""), ""])

#             ws.append(base_row)

#             for col_idx in range(3, len(headers) + 1):
#                 ws.cell(ws.max_row, col_idx).fill = RED_FILL

#             continue

#         # ADDED
#         if row_b and not row_a:

#             base_row.append("ADDED")

#             for col in columns:
#                 base_row.extend(["", row_b.get(col, "")])

#             ws.append(base_row)

#             for col_idx in range(3, len(headers) + 1):
#                 ws.cell(ws.max_row, col_idx).fill = GREEN_FILL

#             continue

#         # CHANGED
#         base_row.append("")

#         data_cells = []
#         changed_flags = []

#         for col in columns:

#             val_a = row_a.get(col, "")
#             val_b = row_b.get(col, "")

#             data_cells.extend([val_a, val_b])

#             changed_flags.append(val_a != val_b)

#         ws.append(base_row + data_cells)

#         excel_row = ws.max_row

#         col_pointer = 3

#         for changed in changed_flags:

#             if changed:
#                 ws.cell(excel_row, col_pointer).fill = YELLOW_FILL
#                 ws.cell(excel_row, col_pointer + 1).fill = YELLOW_FILL

#             col_pointer += 2

#     for col in ws.columns:
#         max_len = max(
#             (len(str(cell.value)) for cell in col if cell.value),
#             default=0,
#         )

#         ws.column_dimensions[col[0].column_letter].width = min(
#             max_len + 2,
#             40,
#         )

#     buf = io.BytesIO()

#     wb.save(buf)

#     return buf.getvalue()